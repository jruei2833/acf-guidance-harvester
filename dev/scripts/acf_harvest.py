#!/usr/bin/env python3
"""
ACF Guidance Document Harvester
================================
Systematically downloads all sub-regulatory guidance documents from the
ACF (Administration for Children and Families) inventory spreadsheet.

Key features:
  - Extracts embedded Excel hyperlinks (not just cell text)
  - Visits each guidance page and scrapes ALL linked documents (PDFs, DOCs, etc.)
  - Creates per-row folders with 4-digit numbering (0001, 0002, ...)
  - Saves metadata JSON for each row
  - Supports batch processing with --start-row / --end-row
  - Wayback Machine fallback for dead links
  - SHA256 checksums for all downloaded files
  - Resume capability (skips already-completed rows)

Usage:
  python acf_harvest.py                          # Process all rows
  python acf_harvest.py --start-row 1 --end-row 100   # Process rows 1-100
  python acf_harvest.py --resume                 # Skip completed rows
  python acf_harvest.py --dry-run                # Parse URLs only, no downloads
"""

import argparse
import hashlib
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
from bs4 import BeautifulSoup

# ─── Configuration ────────────────────────────────────────────────────────────

DEFAULT_INPUT = os.path.join(os.path.dirname(__file__), "..", "input",
                             "Sub-Regulatory_Inventory_for_Programs_12112025_xlsx.xlsx")
DEFAULT_OUTPUT = os.path.join(os.path.dirname(__file__), "..", "output")
DEFAULT_REPORTS = os.path.join(os.path.dirname(__file__), "..", "reports")
DEFAULT_LOGS = os.path.join(os.path.dirname(__file__), "..", "logs")

# Columns of interest (1-indexed for openpyxl)
COL_OFFICE = 1
COL_ISSUE_DATE = 2
COL_TYPE = 3
COL_DOC_NUMBER = 4
COL_TITLE = 5
COL_PAGE_COUNT = 6
COL_DETERMINATION = 7
COL_PROGRAM = 11
COL_URL = 13

# URLs to skip — these appear on pages but are not guidance documents
SKIP_URL_PATTERNS = {
    "vulnerability-disclosure-policy",
    "/privacy-policy",
    "/accessibility",
    "/foia",
    "/disclaimers",
    "/nofear",
    "/plainlanguage",
    "/usa.gov",
    "/digital-strategy",
    "/web-policies-and-important-links",
    "facebook.com",
    "twitter.com",
    "youtube.com",
    "linkedin.com",
    "instagram.com",
}

# File extensions we want to download
DOWNLOAD_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".txt",
    ".rtf", ".ppt", ".pptx", ".zip", ".htm", ".html"
}

# Rate limiting
REQUEST_DELAY = 1.0        # seconds between requests
WAYBACK_DELAY = 2.0        # seconds between Wayback requests
REQUEST_TIMEOUT = 30       # seconds
MAX_RETRIES = 3

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1"
}

# ─── Logging Setup ────────────────────────────────────────────────────────────

def setup_logging(log_dir):
    """Configure logging to both file and console."""
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"harvest_{timestamp}.log")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return log_file


# ─── Utility Functions ────────────────────────────────────────────────────────

def sha256_file(filepath):
    """Compute SHA256 hash of a file."""
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_filename(name, max_len=100):
    """Sanitize a string for use as a filename."""
    name = re.sub(r'[<>:"/\\|?*]', '_', str(name))
    name = re.sub(r'\s+', '_', name)
    name = name.strip('._')
    return name[:max_len] if name else "unnamed"


def fetch_with_retry(url, session, retries=MAX_RETRIES, timeout=REQUEST_TIMEOUT):
    """Fetch a URL with retries and exponential backoff."""
    for attempt in range(retries):
        try:
            resp = session.get(url, timeout=timeout, headers=HEADERS,
                               allow_redirects=True)
            resp.raise_for_status()
            return resp
        except requests.exceptions.RequestException as e:
            if attempt < retries - 1:
                wait = 2 ** (attempt + 1)
                logging.warning(f"  Retry {attempt+1}/{retries} for {url} "
                                f"(waiting {wait}s): {e}")
                time.sleep(wait)
            else:
                logging.error(f"  Failed after {retries} attempts: {url} — {e}")
                return None


def try_alternate_domain(url, session):
    """Try alternate ACF domain if the original fails.
    acf.gov and www.acf.hhs.gov often serve the same content."""
    alt_url = None
    if "acf.gov/" in url and "acf.hhs.gov" not in url:
        alt_url = url.replace("acf.gov/", "www.acf.hhs.gov/", 1)
    elif "www.acf.hhs.gov/" in url:
        alt_url = url.replace("www.acf.hhs.gov/", "acf.gov/", 1)

    if alt_url:
        logging.info(f"  Trying alternate domain: {alt_url}")
        resp = fetch_with_retry(alt_url, session, retries=2)
        if resp:
            return resp
    return None


def try_wayback(url, session):
    """Attempt to retrieve a URL from the Wayback Machine."""
    api_url = f"https://archive.org/wayback/available?url={url}"
    try:
        resp = session.get(api_url, timeout=15, headers=HEADERS)
        data = resp.json()
        snapshot = data.get("archived_snapshots", {}).get("closest", {})
        if snapshot.get("available"):
            wb_url = snapshot["url"]
            logging.info(f"  Wayback fallback found: {wb_url}")
            time.sleep(WAYBACK_DELAY)
            return fetch_with_retry(wb_url, session)
    except Exception as e:
        logging.warning(f"  Wayback lookup failed for {url}: {e}")
    return None


# ─── Excel Parsing ────────────────────────────────────────────────────────────

def extract_row_data(ws, row_idx):
    """Extract all relevant data and URLs from a single Excel row."""
    def cell_val(col):
        v = ws.cell(row=row_idx, column=col).value
        return str(v).strip() if v else ""

    def cell_hyperlink(col):
        cell = ws.cell(row=row_idx, column=col)
        if cell.hyperlink and cell.hyperlink.target:
            return cell.hyperlink.target.strip()
        return None

    # Collect all unique URLs from this row
    urls = []
    seen = set()

    for col in [COL_DOC_NUMBER, COL_TITLE, COL_URL]:
        link = cell_hyperlink(col)
        if link and link not in seen:
            urls.append({"source_column": ws.cell(row=1, column=col).value,
                         "url": link, "type": "hyperlink"})
            seen.add(link)

    # Also check URL column for plain text URLs
    url_text = cell_val(COL_URL)
    if url_text.startswith("http") and url_text not in seen:
        urls.append({"source_column": "URL (text)", "url": url_text,
                      "type": "text"})
        seen.add(url_text)

    return {
        "excel_row": row_idx,
        "office": cell_val(COL_OFFICE),
        "issue_date": cell_val(COL_ISSUE_DATE),
        "doc_type": cell_val(COL_TYPE),
        "doc_number": cell_val(COL_DOC_NUMBER),
        "title": cell_val(COL_TITLE),
        "page_count": cell_val(COL_PAGE_COUNT),
        "determination": cell_val(COL_DETERMINATION),
        "program": cell_val(COL_PROGRAM),
        "urls": urls
    }


# ─── Page Scraping ────────────────────────────────────────────────────────────

def find_downloadable_links(html_content, base_url):
    """Parse an HTML page and find all downloadable document links."""
    soup = BeautifulSoup(html_content, "html.parser")
    links = []
    seen = set()

    for tag in soup.find_all("a", href=True):
        href = tag["href"].strip()
        if not href or href.startswith("#") or href.startswith("mailto:"):
            continue

        full_url = urljoin(base_url, href)
        
        # Skip junk links (footer, nav, social media)
        if any(pattern in full_url.lower() for pattern in SKIP_URL_PATTERNS):
            continue
        
        parsed = urlparse(full_url)
        path_lower = parsed.path.lower()

        # Check if it's a direct file download
        ext = os.path.splitext(path_lower)[1]
        if ext in DOWNLOAD_EXTENSIONS and full_url not in seen:
            link_text = tag.get_text(strip=True)[:200] or os.path.basename(parsed.path)
            links.append({
                "url": full_url,
                "text": link_text,
                "extension": ext
            })
            seen.add(full_url)

    return links


def download_file(url, dest_dir, session, prefix=""):
    """Download a single file and return metadata."""
    resp = fetch_with_retry(url, session)
    if not resp:
        return None

    # Determine filename from URL or Content-Disposition
    filename = ""
    cd = resp.headers.get("Content-Disposition", "")
    if "filename=" in cd:
        match = re.search(r'filename[^;=\n]*=(["\']?)(.+?)\1(;|$)', cd)
        if match:
            filename = match.group(2).strip()

    if not filename:
        parsed = urlparse(url)
        filename = os.path.basename(parsed.path)

    if not filename or filename == "/":
        filename = "document"

    filename = safe_filename(filename)
    if prefix:
        filename = f"{prefix}_{filename}"

    filepath = os.path.join(dest_dir, filename)

    # Handle duplicates
    base, ext = os.path.splitext(filepath)
    counter = 1
    while os.path.exists(filepath):
        filepath = f"{base}_{counter}{ext}"
        counter += 1

    with open(filepath, "wb") as f:
        f.write(resp.content)

    file_hash = sha256_file(filepath)
    file_size = os.path.getsize(filepath)

    logging.info(f"    [OK] Downloaded: {os.path.basename(filepath)} "
                 f"({file_size:,} bytes)")

    return {
        "filename": os.path.basename(filepath),
        "source_url": url,
        "size_bytes": file_size,
        "sha256": file_hash,
        "content_type": resp.headers.get("Content-Type", "unknown")
    }


# ─── Row Processing ──────────────────────────────────────────────────────────

def process_row(row_data, output_dir, session, dry_run=False):
    """Process a single inventory row: visit URLs, scrape, download."""
    row_num = row_data["excel_row"]
    folder_name = f"{row_num:04d}"
    row_dir = os.path.join(output_dir, folder_name)

    result = {
        **row_data,
        "folder": folder_name,
        "status": "pending",
        "files_downloaded": [],
        "errors": [],
        "pages_visited": [],
        "processed_at": datetime.now().isoformat()
    }

    if not row_data["urls"]:
        result["status"] = "no_urls"
        logging.warning(f"Row {row_num:04d}: No URLs found — skipping")
        return result

    if dry_run:
        result["status"] = "dry_run"
        for u in row_data["urls"]:
            logging.info(f"Row {row_num:04d}: [DRY RUN] Would visit {u['url']}")
        return result

    os.makedirs(row_dir, exist_ok=True)

    for url_info in row_data["urls"]:
        url = url_info["url"]
        logging.info(f"Row {row_num:04d}: Visiting {url}")

        time.sleep(REQUEST_DELAY)

        # First try direct fetch
        resp = fetch_with_retry(url, session)

        # Try alternate domain if direct fails
        if not resp:
            resp = try_alternate_domain(url, session)

        # Wayback fallback if both fail
        if not resp:
            logging.info(f"  Trying Wayback Machine for {url}")
            resp = try_wayback(url, session)

        if not resp:
            result["errors"].append({"url": url, "error": "unreachable"})
            continue

        content_type = resp.headers.get("Content-Type", "").lower()
        result["pages_visited"].append({
            "url": url,
            "final_url": resp.url,
            "status_code": resp.status_code,
            "content_type": content_type
        })

        # Case 1: Direct file download (PDF, etc.)
        if any(ct in content_type for ct in
               ["pdf", "msword", "spreadsheet", "zip", "octet-stream"]):
            filepath = os.path.join(row_dir, "direct_download")
            os.makedirs(filepath, exist_ok=True)

            parsed = urlparse(resp.url)
            fname = safe_filename(os.path.basename(parsed.path)) or "document"
            fpath = os.path.join(row_dir, fname)

            # Handle duplicate names
            base, ext = os.path.splitext(fpath)
            counter = 1
            while os.path.exists(fpath):
                fpath = f"{base}_{counter}{ext}"
                counter += 1

            with open(fpath, "wb") as f:
                f.write(resp.content)

            file_meta = {
                "filename": os.path.basename(fpath),
                "source_url": url,
                "size_bytes": os.path.getsize(fpath),
                "sha256": sha256_file(fpath),
                "content_type": content_type,
                "download_type": "direct"
            }
            result["files_downloaded"].append(file_meta)
            logging.info(f"    [OK] Direct download: {file_meta['filename']} "
                         f"({file_meta['size_bytes']:,} bytes)")
            continue

        # Case 2: HTML page — scrape for document links
        if "html" in content_type or "text" in content_type:
            try:
                doc_links = find_downloadable_links(resp.text, resp.url)
                logging.info(f"  Found {len(doc_links)} downloadable links on page")

                for link in doc_links:
                    time.sleep(REQUEST_DELAY)
                    file_meta = download_file(link["url"], row_dir, session)
                    if file_meta:
                        file_meta["link_text"] = link["text"]
                        file_meta["download_type"] = "scraped"
                        result["files_downloaded"].append(file_meta)
                    else:
                        result["errors"].append({
                            "url": link["url"],
                            "error": "download_failed"
                        })

                if not doc_links:
                    # Save the HTML page itself as reference
                    html_path = os.path.join(row_dir, "page_content.html")
                    with open(html_path, "w", encoding="utf-8") as f:
                        f.write(resp.text)
                    logging.info(f"  No doc links found — saved page HTML")

            except Exception as e:
                result["errors"].append({"url": url, "error": str(e)})
                logging.error(f"  Error parsing page: {e}")

    # Determine final status
    if result["files_downloaded"]:
        result["status"] = "success"
    elif result["errors"]:
        result["status"] = "failed"
    else:
        result["status"] = "no_documents"

    # Save row metadata
    meta_path = os.path.join(row_dir, "metadata.json")
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, default=str)

    return result


# ─── Main Execution ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="ACF Guidance Document Harvester")
    parser.add_argument("--input", default=DEFAULT_INPUT,
                        help="Path to Excel inventory file")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help="Output directory for downloaded files")
    parser.add_argument("--reports", default=DEFAULT_REPORTS,
                        help="Directory for summary reports")
    parser.add_argument("--logs", default=DEFAULT_LOGS,
                        help="Directory for log files")
    parser.add_argument("--start-row", type=int, default=1,
                        help="First row to process (1-indexed, relative to data)")
    parser.add_argument("--end-row", type=int, default=None,
                        help="Last row to process (inclusive)")
    parser.add_argument("--resume", action="store_true",
                        help="Skip rows that already have output folders")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse and log URLs without downloading")
    args = parser.parse_args()

    # Resolve paths
    input_path = os.path.abspath(args.input)
    output_dir = os.path.abspath(args.output)
    reports_dir = os.path.abspath(args.reports)
    logs_dir = os.path.abspath(args.logs)

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(reports_dir, exist_ok=True)

    log_file = setup_logging(logs_dir)
    logging.info(f"ACF Guidance Harvester starting")
    logging.info(f"  Input:   {input_path}")
    logging.info(f"  Output:  {output_dir}")
    logging.info(f"  Reports: {reports_dir}")
    logging.info(f"  Log:     {log_file}")

    # Load workbook
    logging.info("Loading Excel workbook...")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
    total_rows = ws.max_row - 1  # minus header
    logging.info(f"  Found {total_rows} data rows in sheet '{ws.title}'")

    # Determine row range
    start = max(1, args.start_row)
    end = min(total_rows, args.end_row) if args.end_row else total_rows
    logging.info(f"  Processing rows {start} to {end} "
                 f"({end - start + 1} rows)")

    # Initialize session
    session = requests.Session()
    session.headers.update(HEADERS)

    # Track results
    results = []
    stats = {"success": 0, "failed": 0, "no_urls": 0,
             "no_documents": 0, "skipped": 0, "dry_run": 0,
             "total_files": 0}

    start_time = datetime.now()

    for i in range(start, end + 1):
        excel_row = i + 1  # +1 for header row
        folder_name = f"{i:04d}"

        # Resume check
        if args.resume:
            row_dir = os.path.join(output_dir, folder_name)
            meta_file = os.path.join(row_dir, "metadata.json")
            if os.path.exists(meta_file):
                logging.info(f"Row {i:04d}: Skipping (already processed)")
                stats["skipped"] += 1
                continue

        # Extract and process
        row_data = extract_row_data(ws, excel_row)
        result = process_row(row_data, output_dir, session,
                             dry_run=args.dry_run)
        results.append(result)

        # Update stats
        status = result["status"]
        if status in stats:
            stats[status] += 1
        stats["total_files"] += len(result.get("files_downloaded", []))

        # Progress update every 50 rows
        processed = i - start + 1
        if processed % 50 == 0:
            elapsed = (datetime.now() - start_time).total_seconds()
            rate = processed / elapsed * 3600 if elapsed > 0 else 0
            logging.info(f"── Progress: {processed}/{end - start + 1} rows "
                         f"({rate:.0f} rows/hr) ──")

    # Final summary
    elapsed = (datetime.now() - start_time).total_seconds()
    logging.info("=" * 60)
    logging.info("HARVEST COMPLETE")
    logging.info(f"  Duration: {elapsed/60:.1f} minutes")
    logging.info(f"  Rows processed: {len(results)}")
    logging.info(f"  Success: {stats['success']}")
    logging.info(f"  Failed: {stats['failed']}")
    logging.info(f"  No URLs: {stats['no_urls']}")
    logging.info(f"  No documents found: {stats['no_documents']}")
    logging.info(f"  Skipped (resume): {stats['skipped']}")
    logging.info(f"  Total files downloaded: {stats['total_files']}")
    logging.info("=" * 60)

    # Save summary report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report = {
        "run_timestamp": timestamp,
        "duration_seconds": elapsed,
        "row_range": {"start": start, "end": end},
        "stats": stats,
        "results": results
    }

    report_path = os.path.join(reports_dir, f"harvest_report_{timestamp}.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str)
    logging.info(f"Report saved: {report_path}")

    # Also save a CSV summary for quick reference
    csv_path = os.path.join(reports_dir, f"harvest_summary_{timestamp}.csv")
    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("row,folder,office,doc_number,title,status,files_count,errors_count\n")
        for r in results:
            f.write(f'{r["excel_row"]},{r["folder"]},{r["office"]},'
                    f'"{r["doc_number"]}","{r["title"][:80]}",'
                    f'{r["status"]},{len(r["files_downloaded"])},'
                    f'{len(r["errors"])}\n')
    logging.info(f"CSV summary saved: {csv_path}")

    wb.close()
    return 0 if stats["failed"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
