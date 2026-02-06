#!/usr/bin/env python3
"""
build_linked_inventory.py - Take the original ACF inventory Excel and add
a "File Location" column linking to each document in the organized folder.

For missing documents, inserts a note explaining why.

Output: catalog/ACF_Inventory_Linked.xlsx
"""

import json
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# --- Configuration ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
ORGANIZED_DIR = os.path.join(BASE_DIR, "organized")
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")
INPUT_DIR = os.path.join(BASE_DIR, "input")

# Missing document reasons (from the manifest)
MISSING_REASONS = {
    "0835": "Server returns 403 Forbidden; page blocked at server level",
    "0838": "Server returns 403 Forbidden; redirect URL also blocked",
    "0840": "Server returns 403 Forbidden on both live site and Wayback Machine",
    "1212": "Page returns 404 Not Found; legacy URL no longer resolves",
    "1227": "Server returns 403 Forbidden",
    "1275": "No match found on any ACF index page; no Wayback snapshots",
    "1292": "Page returns 404 Not Found; legacy URL with misspelling no longer resolves",
    "1298": "Server returns 403 Forbidden on both live and alternate domains",
    "2291": "Server returns 403 Forbidden despite being listed on ORR DCL index",
    "2392": "Page returns 404 Not Found; /resource/ path no longer exists",
    "2394": "Server returns 403 Forbidden on redirected URL",
    "2396": "Page returns 404 Not Found",
    "2397": "Page returns 404 Not Found",
    "2399": "Server returns 403 Forbidden on redirected URL",
    "2401": "Page returns 404 Not Found",
    "2402": "Page returns 404 Not Found",
    "2403": "Page returns 404 Not Found",
    "2407": "Page returns 404 Not Found",
    "2410": "Server returns 403 Forbidden on redirected URL",
    "2417": "Server returns 403 Forbidden on redirected URL",
    "2422": "Page returns 404 Not Found",
    "2426": "Server returns 403 Forbidden on redirected URL",
    "2546": "Server returns 403 Forbidden; State Letters appear systematically blocked",
    "2547": "Server returns 403 Forbidden",
    "2553": "Server returns 403 Forbidden",
    "2584": "Server returns 403 Forbidden",
    "2647": "Server returns 403 Forbidden",
    "2653": "Page returns 404 Not Found; archive path does not resolve",
    "2659": "Server returns 403 Forbidden",
    "2682": "Page returns 404 Not Found",
    "2694": "Page returns 404 Not Found",
    "2704": "Page returns 404 Not Found",
    "2707": "Page returns 404 Not Found",
    "2721": "Page returns 404 Not Found",
    "2758": "No URL available; no match found on any index page",
    "3365": "No match found; headstart.gov content not indexed",
    "3366": "No match found; headstart.gov content not indexed",
}


def find_primary_file(folder_id):
    """Find the primary file for a given row in the organized structure."""
    meta_path = os.path.join(OUTPUT_DIR, folder_id, "metadata.json")
    if not os.path.isfile(meta_path):
        return None, None
    
    try:
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
    except:
        return None, None
    
    if meta.get("status") != "success":
        return None, None
    
    office = meta.get("office", "")
    doc_type_raw = meta.get("doc_type", "")
    
    # Find this row's folder in organized/
    # Walk the organized dir to find the folder matching this row ID
    for root, dirs, files in os.walk(ORGANIZED_DIR):
        if os.path.basename(root) == folder_id:
            # Found it - get the primary file (first PDF, or first file)
            pdfs = [f for f in files if f.lower().endswith(".pdf")]
            if pdfs:
                # Primary = the one without "_original" or "_Attachment"
                for pdf in pdfs:
                    if "_original" not in pdf and "_Attachment" not in pdf:
                        rel_path = os.path.relpath(os.path.join(root, pdf), CATALOG_DIR)
                        folder_rel = os.path.relpath(root, CATALOG_DIR)
                        return rel_path, folder_rel
                # Fallback to first PDF
                rel_path = os.path.relpath(os.path.join(root, pdfs[0]), CATALOG_DIR)
                folder_rel = os.path.relpath(root, CATALOG_DIR)
                return rel_path, folder_rel
            elif files:
                # No PDFs, use first non-metadata file
                for f in files:
                    if f != "metadata.json":
                        rel_path = os.path.relpath(os.path.join(root, f), CATALOG_DIR)
                        folder_rel = os.path.relpath(root, CATALOG_DIR)
                        return rel_path, folder_rel
    
    return None, None


def main():
    print("=" * 60)
    print("BUILD LINKED INVENTORY")
    print("=" * 60)
    
    os.makedirs(CATALOG_DIR, exist_ok=True)
    
    # Find original inventory
    input_files = [f for f in os.listdir(INPUT_DIR) if f.endswith(".xlsx")]
    if not input_files:
        print("ERROR: No .xlsx file found in input/")
        sys.exit(1)
    
    src_path = os.path.join(INPUT_DIR, input_files[0])
    print(f"\nSource: {input_files[0]}")
    
    # Load original workbook (read-only to get data)
    print("Loading original inventory...")
    src_wb = openpyxl.load_workbook(src_path, data_only=True)
    src_ws = src_wb.active
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACF Guidance Inventory"
    
    # Styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    missing_font = Font(name="Arial", size=10, color="CC0000", italic=True)
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    # Build index of all row folders and their primary files
    print("Indexing organized documents...")
    file_index = {}  # folder_id -> (file_rel_path, folder_rel_path)
    
    folders = sorted([d for d in os.listdir(OUTPUT_DIR)
                      if os.path.isdir(os.path.join(OUTPUT_DIR, d)) and d.isdigit()])
    
    for folder_id in folders:
        file_path, folder_path = find_primary_file(folder_id)
        if file_path:
            file_index[folder_id] = (file_path, folder_path)
    
    print(f"  Indexed {len(file_index)} documents")
    
    # Copy original data + add File Location column
    print("Building linked inventory...")
    
    max_col = src_ws.max_column
    new_col = max_col + 1  # File Location column
    
    row_count = 0
    linked = 0
    missing_noted = 0
    
    for row in src_ws.iter_rows():
        row_count += 1
        
        # Copy all original cells
        for col_idx, cell in enumerate(row, 1):
            new_cell = ws.cell(row=row_count, column=col_idx, value=cell.value)
            new_cell.border = thin_border
            
            if row_count == 1:
                new_cell.font = header_font
                new_cell.fill = header_fill
                new_cell.alignment = header_align
            else:
                new_cell.font = data_font
                if row_count % 2 == 0:
                    new_cell.fill = alt_fill
        
        # Add File Location column
        loc_cell = ws.cell(row=row_count, column=new_col)
        loc_cell.border = thin_border
        
        if row_count == 1:
            # Header
            loc_cell.value = "File Location"
            loc_cell.font = header_font
            loc_cell.fill = header_fill
            loc_cell.alignment = header_align
        else:
            # Data row - folder_id is row_count padded to 4 digits
            # Excel row 2 = folder 0002, row 3 = folder 0003, etc.
            folder_id = f"{row_count:04d}"
            
            if folder_id in file_index:
                file_path, folder_path = file_index[folder_id]
                link_path = file_path.replace("\\", "/")
                loc_cell.value = "Open Document"
                loc_cell.hyperlink = link_path
                loc_cell.font = link_font
                if row_count % 2 == 0:
                    loc_cell.fill = alt_fill
                linked += 1
            elif folder_id in MISSING_REASONS:
                loc_cell.value = f"UNAVAILABLE: {MISSING_REASONS[folder_id]}"
                loc_cell.font = missing_font
                if row_count % 2 == 0:
                    loc_cell.fill = alt_fill
                missing_noted += 1
            else:
                # Check if it's a failed row not in our explicit list
                meta_path = os.path.join(OUTPUT_DIR, folder_id, "metadata.json")
                if os.path.isfile(meta_path):
                    try:
                        with open(meta_path) as f:
                            meta = json.load(f)
                        if meta.get("status") == "failed":
                            errors = meta.get("errors", [])
                            if errors:
                                reason = errors[0].get("error", "unknown error")
                            else:
                                reason = "download failed"
                            loc_cell.value = f"UNAVAILABLE: {reason}"
                            loc_cell.font = missing_font
                            missing_noted += 1
                        else:
                            loc_cell.value = ""
                    except:
                        loc_cell.value = ""
                else:
                    loc_cell.value = ""
                
                if row_count % 2 == 0:
                    loc_cell.fill = alt_fill
    
    # Set column widths
    for col in range(1, new_col):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions[get_column_letter(new_col)].width = 45
    
    # Freeze and filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(new_col)}1"
    
    # Save
    output_path = os.path.join(CATALOG_DIR, "ACF_Inventory_Linked.xlsx")
    wb.save(output_path)
    
    print(f"\n{'=' * 60}")
    print(f"COMPLETE")
    print(f"{'=' * 60}")
    print(f"  Rows processed: {row_count - 1}")
    print(f"  Documents linked: {linked}")
    print(f"  Missing noted: {missing_noted}")
    print(f"  Output: {output_path}")
    print(f"\n  Place this file alongside the 'organized/' folder so hyperlinks resolve.")


if __name__ == "__main__":
    main()
