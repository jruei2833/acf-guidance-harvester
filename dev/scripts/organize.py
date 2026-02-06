#!/usr/bin/env python3
"""
organize.py - Reorganize harvested ACF guidance into a clean folder structure
with a master Excel catalog for downstream processing.

Creates:
  organized/
    {Office}/
      {Doc Type}/
        {Doc ID}/
          {Doc ID}.pdf                         <- primary (PDF version)
          {Doc ID}_original.docx               <- primary (original format)
          {Doc ID}_CFS101_Part_I_Form.pdf      <- attachment (PDF conversion)
          {Doc ID}_CFS101_Part_I_Form.xls      <- attachment (original format)
          ...

  catalog/
    ACF_Guidance_Master_Catalog.xlsx
      Tab 1: "Master Catalog" - full index with hyperlinks to files
      Tab 2: "Summary" - per-office coverage stats
      Tab 3: "Original Inventory" - copy of the source Excel data
"""

import json
import os
import re
import sys
import shutil
import time
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# --- Configuration ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
ORGANIZED_DIR = os.path.join(BASE_DIR, "organized")
CATALOG_DIR = os.path.join(BASE_DIR, "catalog")
REPORT_DIR = os.path.join(BASE_DIR, "reports")
INPUT_DIR = os.path.join(BASE_DIR, "input")

# --- Helpers ---

def sanitize_folder_name(name):
    if not name:
        return "Unknown"
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', ' ', name).strip()
    name = name.strip('. ')
    if len(name) > 80:
        name = name[:80].strip()
    return name


def sanitize_filename(name):
    if not name:
        return "document"
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'%[0-9A-Fa-f]{2}', '_', name)
    name = re.sub(r'\s+', '_', name).strip('_. ')
    if len(name) > 80:
        name = name[:80].strip('_')
    return name


def clean_doc_type(doc_type):
    if not doc_type:
        return "Other"
    
    type_map = {
        "action transmittal": "Action Transmittals",
        "dear colleague letter": "Dear Colleague Letters",
        "information memorandum": "Information Memoranda",
        "program instruction": "Program Instructions",
        "policy announcement": "Policy Announcements",
        "policy interpretation question": "Policy Interpretation Questions",
        "child welfare policy manual": "Child Welfare Policy Manual",
        "technical bulletin": "Technical Bulletins",
        "frequently asked questions": "FAQs",
        "state letter": "State Letters",
        "grant letter": "Grant Letters",
        "formula grant allocation": "Formula Grant Allocations",
        "tribal consultation": "Tribal Consultations",
        "regulatory action": "Regulatory Actions",
        "report": "Reports",
    }
    
    dt_lower = doc_type.lower().strip()
    
    for key, value in type_map.items():
        if key in dt_lower:
            return value
    
    if "/" in doc_type:
        parts = [p.strip() for p in doc_type.split("/")]
        for part in parts:
            clean = clean_doc_type(part)
            if clean != "Other":
                return clean
    
    return sanitize_folder_name(doc_type)


def clean_doc_id(doc_number, folder_num, office):
    if doc_number and doc_number.strip():
        doc_id = doc_number.strip()
        for prefix in ["CCDF-ACF-", "ACF-", "ACYF-"]:
            if doc_id.upper().startswith(prefix):
                doc_id = doc_id[len(prefix):]
        return sanitize_folder_name(doc_id)
    else:
        return f"ROW-{folder_num}"


def get_pdf_version(filename, folder_path, conversions):
    if filename.lower().endswith(".pdf"):
        return filename
    if conversions:
        for c in conversions:
            if c.get("original_file") == filename:
                pdf_name = c.get("pdf_file", "")
                if os.path.isfile(os.path.join(folder_path, pdf_name)):
                    return pdf_name
    stem = os.path.splitext(filename)[0]
    pdf_name = f"{stem}.pdf"
    if os.path.isfile(os.path.join(folder_path, pdf_name)):
        return pdf_name
    return None


def identify_primary_file(files_downloaded, conversions=None):
    if not files_downloaded:
        return None, []
    
    if len(files_downloaded) == 1:
        return files_downloaded[0], []
    
    pdfs = [f for f in files_downloaded if f.get("filename", "").lower().endswith(".pdf")]
    
    if pdfs:
        for pdf in pdfs:
            fname = pdf.get("filename", "").lower()
            link = pdf.get("link_text", "").lower()
            if any(kw in fname or kw in link for kw in 
                   ["attachment", "form", "appendix", "addendum", "worksheet", "template", "cfs101", "cfs-101"]):
                continue
            return pdf, [f for f in files_downloaded if f != pdf]
        return pdfs[0], [f for f in files_downloaded if f != pdfs[0]]
    
    return files_downloaded[0], files_downloaded[1:]


def copy_file_safe(src, dst):
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.copy2(src, dst)


# --- Main ---

def main():
    start_time = time.time()
    
    print("=" * 70)
    print("ACF GUIDANCE — ORGANIZE & CATALOG")
    print("=" * 70)
    
    if os.path.exists(ORGANIZED_DIR):
        print(f"\nRemoving existing organized directory...")
        shutil.rmtree(ORGANIZED_DIR)
    os.makedirs(ORGANIZED_DIR, exist_ok=True)
    os.makedirs(CATALOG_DIR, exist_ok=True)
    os.makedirs(REPORT_DIR, exist_ok=True)
    
    # Phase 1: Load all metadata
    print("\nPhase 1: Loading metadata...")
    
    folders = sorted([d for d in os.listdir(OUTPUT_DIR)
                      if os.path.isdir(os.path.join(OUTPUT_DIR, d)) and d.isdigit()])
    
    all_rows = []
    for folder in folders:
        meta_path = os.path.join(OUTPUT_DIR, folder, "metadata.json")
        if not os.path.isfile(meta_path):
            continue
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                meta = json.load(f)
            meta["_folder"] = folder
            all_rows.append(meta)
        except Exception as e:
            print(f"  Warning: Could not read {meta_path}: {e}")
    
    print(f"  Loaded {len(all_rows)} rows")
    
    # Phase 2: Organize files
    print("\nPhase 2: Organizing files into clean structure...")
    
    catalog_entries = []
    files_copied = 0
    folders_created = 0
    
    for i, meta in enumerate(all_rows):
        if (i + 1) % 500 == 0:
            print(f"  [{i+1}/{len(all_rows)}] Processing...")
        
        folder = meta["_folder"]
        folder_path = os.path.join(OUTPUT_DIR, folder)
        
        office = meta.get("office", "Unknown")
        doc_type_raw = meta.get("doc_type", "Other")
        doc_number = meta.get("doc_number", "")
        title = meta.get("title", "")
        issue_date = meta.get("issue_date", "")
        status = meta.get("status", "unknown")
        excel_row = meta.get("excel_row", "")
        conversions = meta.get("conversions", [])
        files_downloaded = meta.get("files_downloaded", [])
        
        doc_type_folder = clean_doc_type(doc_type_raw)
        doc_id = clean_doc_id(doc_number, folder, office)
        
        # Use row number as folder name for direct Excel correlation
        dest_dir = os.path.join(ORGANIZED_DIR, office, doc_type_folder, folder)
        rel_dest = os.path.relpath(dest_dir, CATALOG_DIR) if dest_dir else ""
        
        entry = {
            "row_id": folder,
            "excel_row": excel_row,
            "office": office,
            "doc_type": doc_type_raw,
            "doc_type_clean": doc_type_folder,
            "doc_number": doc_number,
            "doc_id": doc_id,
            "title": title,
            "issue_date": issue_date,
            "status": status,
            "primary_file": "",
            "primary_file_link": "",
            "num_files": len(files_downloaded),
            "num_attachments": 0,
            "all_files": "",
            "folder_path": "",
            "folder_link": rel_dest,
            "source_url": "",
        }
        
        urls = meta.get("urls", [])
        if urls:
            entry["source_url"] = urls[0].get("url", "")
        
        if status != "success" or not files_downloaded:
            catalog_entries.append(entry)
            continue
        
        primary, attachments = identify_primary_file(files_downloaded, conversions)
        
        if not primary:
            catalog_entries.append(entry)
            continue
        
        os.makedirs(dest_dir, exist_ok=True)
        folders_created += 1
        
        file_list = []
        primary_copied_name = ""
        
        # --- Copy primary document ---
        primary_filename = primary.get("filename", "")
        primary_ext = os.path.splitext(primary_filename)[1].lower()
        primary_pdf = get_pdf_version(primary_filename, folder_path, conversions)
        
        if primary_pdf:
            src = os.path.join(folder_path, primary_pdf)
            dst_name = f"{doc_id}.pdf"
            dst = os.path.join(dest_dir, dst_name)
            if os.path.isfile(src):
                copy_file_safe(src, dst)
                files_copied += 1
                file_list.append(dst_name)
                primary_copied_name = dst_name
        
        if primary_ext != ".pdf":
            src = os.path.join(folder_path, primary_filename)
            if os.path.isfile(src):
                dst_name = f"{doc_id}_original{primary_ext}"
                dst = os.path.join(dest_dir, dst_name)
                copy_file_safe(src, dst)
                files_copied += 1
                file_list.append(dst_name)
                if not primary_copied_name:
                    primary_copied_name = dst_name
        elif not primary_pdf:
            src = os.path.join(folder_path, primary_filename)
            if os.path.isfile(src):
                dst_name = f"{doc_id}.pdf"
                dst = os.path.join(dest_dir, dst_name)
                copy_file_safe(src, dst)
                files_copied += 1
                file_list.append(dst_name)
                primary_copied_name = dst_name
        
        # --- Copy attachments ---
        for att_idx, att in enumerate(attachments, 1):
            att_filename = att.get("filename", "")
            att_ext = os.path.splitext(att_filename)[1].lower()
            link_text = att.get("link_text", "")
            
            if link_text:
                att_label = sanitize_filename(os.path.splitext(link_text)[0])
                # Strip doc_id prefix if link_text already contains it
                doc_id_clean = sanitize_filename(doc_id)
                if att_label.startswith(doc_id_clean + "_"):
                    att_label = att_label[len(doc_id_clean) + 1:]
                elif att_label.startswith(doc_id_clean):
                    att_label = att_label[len(doc_id_clean):].lstrip("_")
                if not att_label:
                    att_label = f"Attachment_{att_idx}"
            else:
                att_label = f"Attachment_{att_idx}"
            
            att_pdf = get_pdf_version(att_filename, folder_path, conversions)
            if att_pdf:
                src = os.path.join(folder_path, att_pdf)
                if os.path.isfile(src):
                    dst_name = f"{doc_id}_{att_label}.pdf"
                    dst = os.path.join(dest_dir, dst_name)
                    copy_file_safe(src, dst)
                    files_copied += 1
                    file_list.append(dst_name)
            
            src = os.path.join(folder_path, att_filename)
            if os.path.isfile(src):
                if att_ext.lower() != ".pdf":
                    dst_name = f"{doc_id}_{att_label}{att_ext}"
                    dst = os.path.join(dest_dir, dst_name)
                    copy_file_safe(src, dst)
                    files_copied += 1
                    file_list.append(dst_name)
                elif not att_pdf:
                    dst_name = f"{doc_id}_{att_label}.pdf"
                    dst = os.path.join(dest_dir, dst_name)
                    copy_file_safe(src, dst)
                    files_copied += 1
                    file_list.append(dst_name)
        
        entry["primary_file"] = primary_copied_name
        entry["num_attachments"] = len(attachments)
        entry["all_files"] = "; ".join(file_list)
        entry["folder_path"] = os.path.relpath(dest_dir, ORGANIZED_DIR)
        
        if primary_copied_name:
            primary_full = os.path.join(dest_dir, primary_copied_name)
            entry["primary_file_link"] = os.path.relpath(primary_full, CATALOG_DIR)
        
        catalog_entries.append(entry)
    
    print(f"  Folders created: {folders_created}")
    print(f"  Files copied: {files_copied}")
    
    # Phase 3: Generate Master Catalog Excel
    print("\nPhase 3: Generating Master Catalog Excel...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Catalog"
    
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    success_font = Font(name="Arial", size=10, color="006600")
    failed_font = Font(name="Arial", size=10, color="CC0000")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    
    headers = [
        ("Row ID", 8),
        ("Office", 8),
        ("Doc Type", 25),
        ("Doc Number", 18),
        ("Doc ID", 18),
        ("Title", 45),
        ("Issue Date", 12),
        ("Status", 10),
        ("Primary File", 30),
        ("Link to File", 15),
        ("# Files", 8),
        ("# Attachments", 12),
        ("All Files", 60),
        ("Folder Path", 40),
        ("Link to Folder", 15),
        ("Source URL", 50),
    ]
    
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    
    for row_idx, entry in enumerate(catalog_entries, 2):
        values = [
            entry["row_id"],
            entry["office"],
            entry["doc_type"],
            entry["doc_number"],
            entry["doc_id"],
            entry["title"],
            entry["issue_date"][:10] if entry["issue_date"] else "",
            entry["status"],
            entry["primary_file"],
            "",
            entry["num_files"],
            entry["num_attachments"],
            entry["all_files"],
            entry["folder_path"],
            "",
            entry["source_url"],
        ]
        
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [6, 13]))
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            if col_idx == 8:
                if value == "success":
                    cell.font = success_font
                elif value == "failed":
                    cell.font = failed_font
        
        if entry["primary_file_link"]:
            link_cell = ws.cell(row=row_idx, column=10)
            link_path = entry["primary_file_link"].replace("\\", "/")
            link_cell.value = "Open File"
            link_cell.hyperlink = link_path
            link_cell.font = link_font
            link_cell.border = thin_border
            if row_idx % 2 == 0:
                link_cell.fill = alt_fill
        
        if entry["folder_link"] and entry["status"] == "success":
            folder_cell = ws.cell(row=row_idx, column=15)
            folder_path = entry["folder_link"].replace("\\", "/")
            folder_cell.value = "Open Folder"
            folder_cell.hyperlink = folder_path
            folder_cell.font = link_font
            folder_cell.border = thin_border
            if row_idx % 2 == 0:
                folder_cell.fill = alt_fill
    
    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15
    ws2.column_dimensions["D"].width = 15
    
    summary_headers = ["Category", "Total", "Success", "Coverage"]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    offices = sorted(set(e["office"] for e in catalog_entries))
    for row_idx, office in enumerate(offices, 2):
        office_rows = [e for e in catalog_entries if e["office"] == office]
        success = len([e for e in office_rows if e["status"] == "success"])
        total = len(office_rows)
        pct = f"{success/total*100:.1f}%" if total > 0 else "0%"
        ws2.cell(row=row_idx, column=1, value=office).font = data_font
        ws2.cell(row=row_idx, column=2, value=total).font = data_font
        ws2.cell(row=row_idx, column=3, value=success).font = data_font
        ws2.cell(row=row_idx, column=4, value=pct).font = data_font
        for c in range(1, 5):
            ws2.cell(row=row_idx, column=c).border = thin_border
            if row_idx % 2 == 0:
                ws2.cell(row=row_idx, column=c).fill = alt_fill
    
    total_row = len(offices) + 2
    total_all = len(catalog_entries)
    total_success = len([e for e in catalog_entries if e["status"] == "success"])
    bold_font = Font(name="Arial", size=10, bold=True)
    ws2.cell(row=total_row, column=1, value="TOTAL").font = bold_font
    ws2.cell(row=total_row, column=2, value=total_all).font = bold_font
    ws2.cell(row=total_row, column=3, value=total_success).font = bold_font
    ws2.cell(row=total_row, column=4, value=f"{total_success/total_all*100:.1f}%").font = bold_font
    for c in range(1, 5):
        ws2.cell(row=total_row, column=c).border = thin_border
    
    # --- Original Inventory tab ---
    print("  Loading original inventory spreadsheet...")
    
    input_files = [f for f in os.listdir(INPUT_DIR) if f.endswith(".xlsx")]
    if input_files:
        src_wb_path = os.path.join(INPUT_DIR, input_files[0])
        try:
            src_wb = openpyxl.load_workbook(src_wb_path, read_only=True, data_only=True)
            src_ws = src_wb.active
            
            ws3 = wb.create_sheet("Original Inventory")
            
            print(f"  Copying from: {input_files[0]}")
            row_count = 0
            for row in src_ws.iter_rows():
                row_count += 1
                for col_idx, cell in enumerate(row, 1):
                    new_cell = ws3.cell(row=row_count, column=col_idx, value=cell.value)
                    if row_count == 1:
                        new_cell.font = header_font
                        new_cell.fill = header_fill
                        new_cell.alignment = header_align
                    else:
                        new_cell.font = data_font
                        if row_count % 2 == 0:
                            new_cell.fill = alt_fill
                    new_cell.border = thin_border
            
            for col in range(1, ws3.max_column + 1):
                ws3.column_dimensions[get_column_letter(col)].width = 18
            
            ws3.freeze_panes = "A2"
            print(f"  Copied {row_count} rows from original inventory")
            
            src_wb.close()
        except Exception as e:
            print(f"  Warning: Could not load original inventory: {e}")
    else:
        print("  Warning: No .xlsx file found in input/ directory")
    
    # Save
    catalog_path = os.path.join(CATALOG_DIR, "ACF_Guidance_Master_Catalog.xlsx")
    wb.save(catalog_path)
    print(f"  Catalog saved: {catalog_path}")
    
    # Phase 4: Summary
    elapsed = time.time() - start_time
    
    office_counts = {}
    for entry in catalog_entries:
        if entry["status"] == "success" and entry["folder_path"]:
            office = entry["office"]
            if office not in office_counts:
                office_counts[office] = 0
            office_counts[office] += 1
    
    total_size = 0
    for root, dirs, files in os.walk(ORGANIZED_DIR):
        for f in files:
            total_size += os.path.getsize(os.path.join(root, f))
    
    print("\n" + "=" * 70)
    print(f"ORGANIZATION COMPLETE ({elapsed:.1f} seconds)")
    print("=" * 70)
    print(f"  Total rows cataloged: {len(catalog_entries)}")
    print(f"  Documents organized: {folders_created}")
    print(f"  Files copied: {files_copied}")
    print(f"  Total organized size: {total_size / (1024**3):.1f} GB")
    print(f"  Failed/unavailable: {len([e for e in catalog_entries if e['status'] != 'success'])}")
    print(f"\n  By office:")
    for office in sorted(office_counts):
        print(f"    {office}: {office_counts[office]} documents")
    print(f"\n  Organized folder: {ORGANIZED_DIR}")
    print(f"  Master catalog: {catalog_path}")
    
    report = {
        "timestamp": datetime.now().isoformat(),
        "elapsed_seconds": elapsed,
        "total_cataloged": len(catalog_entries),
        "documents_organized": folders_created,
        "files_copied": files_copied,
        "total_size_bytes": total_size,
        "by_office": office_counts,
    }
    report_path = os.path.join(REPORT_DIR, f"organize_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2)
    print(f"  Report saved: {report_path}")


if __name__ == "__main__":
    main()
