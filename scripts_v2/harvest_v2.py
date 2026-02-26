#!/usr/bin/env python3
"""
ACF Guidance Harvester v2
==========================
Full reharvest with upgraded recovery methods.

Key improvements over v1:
  1. Content validation gate — every download is checked for real content
  2. Source priority chain — Portal → Direct URL → Wayback (validated)
  3. HTML quality detection — catches DOJ blocks, Wayback junk, empty shells
  4. Smart PDF detection — catches HTML files masquerading as PDFs
  5. Clean Playwright rendering — captures visible content area only

Usage:
  python harvest_v2.py                          # Full run, all rows
  python harvest_v2.py --start 1 --end 500      # Specific range
  python harvest_v2.py --resume                  # Skip completed rows
  python harvest_v2.py --dry-run                 # Preview only
  python harvest_v2.py --validate-only           # Re-validate existing output_v2

Requirements:
  pip install openpyxl requests beautifulsoup4 playwright
  playwright install chromium
"""

import os
import sys
import json
import re
import time
import hashlib
import argparse
import logging
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse, urljoin

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl

# Import our validation module
from validate import validate_content, detect_file_type, ValidationResult

# ── Configuration ────────────────────────────────────────────────────────

# Paths — adjust if your repo structure differs
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
INPUT_EXCEL = os.path.join(BASE_DIR, "input", "Sub-Regulatory Inventory for Programs_12112025.xlsx.xlsx")
CATALOG_EXCEL = os.path.join(BASE_DIR, "catalog", "ACF_Guidance_Master_Catalog.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output_v2")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")

# If running from the scripts_v2 folder directly, also check parent
if not os.path.exists(os.path.join(BASE_DIR, "input")):
    # Try one level up
    BASE_DIR = os.path.dirname(BASE_DIR)
    INPUT_EXCEL = os.path.join(BASE_DIR, "input", "Sub-Regulatory Inventory for Programs_12112025.xlsx.xlsx")
    CATALOG_EXCEL = os.path.join(BASE_DIR, "catalog", "ACF_Guidance_Master_Catalog.xlsx")
    OUTPUT_DIR = os.path.join(BASE_DIR, "output_v2")
    REPORTS_DIR = os.path.join(BASE_DIR, "reports")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,"
              "application/pdf,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
}

# Rate limiting
REQUEST_DELAY = 1.5          # seconds between requests
WAYBACK_DELAY = 2.0          # be polite to archive.org
PLAYWRIGHT_DELAY = 2.0       # between browser page loads

# Validation thresholds
MIN_TEXT_CHARS = 500          # minimum visible text for HTML to be considered real
MIN_FILE_SIZE = 200           # minimum bytes for any file

# Logging
LOG_FORMAT = "%(asctime)s [%(levelname)s] %(message)s"
LOG_DATE = "%H:%M:%S"

# ── Setup ────────────────────────────────────────────────────────────────

def setup_logging(verbose=False):
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=level, format=LOG_FORMAT, datefmt=LOG_DATE)
    return logging.getLogger("acf_v2")


def setup_session():
    """Create a requests session with retry logic."""
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["HEAD", "GET"],
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    session.headers.update(HEADERS)
    return session


def sha256_file(filepath):
    h = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def safe_filename(name, max_len=100):
    """Sanitize a string for use as a filename."""
    name = re.sub(r'[<>:"/\\|?*]', '_', name)
    name = re.sub(r'\s+', '_', name).strip('_.')
    return name[:max_len] if name else "document"


# ── Excel reading ────────────────────────────────────────────────────────

def extract_urls_from_row(row, col_indices):
    """
    Extract all URLs from a row, checking both hyperlinks and cell values.
    Returns list of unique URLs.
    """
    urls = []
    seen = set()

    for col_idx in col_indices:
        cell = row[col_idx] if col_idx < len(row) else None
        if cell is None:
            continue

        # Check hyperlink first (the v1 key discovery)
        if hasattr(cell, 'hyperlink') and cell.hyperlink and cell.hyperlink.target:
            url = cell.hyperlink.target.strip()
            if url and url not in seen and url.startswith("http"):
                urls.append(url)
                seen.add(url)

        # Also check cell value for URLs
        val = str(cell.value or "").strip()
        if val.startswith("http") and val not in seen:
            urls.append(val)
            seen.add(val)

        # Check for URLs embedded in text
        url_pattern = re.findall(r'https?://[^\s<>"\']+', val)
        for u in url_pattern:
            u = u.rstrip('.,;:)')
            if u not in seen:
                urls.append(u)
                seen.add(u)

    return urls


def load_inventory(excel_path, start_row=1, end_row=None):
    """
    Load the Excel inventory and extract row data with URLs.
    Returns list of dicts with row_id, office, doc_number, title, urls, etc.
    """
    logger = logging.getLogger("acf_v2")
    logger.info(f"Loading inventory from: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb.active

    # Find column indices by header names
    headers = {}
    for idx, cell in enumerate(ws[1]):
        val = str(cell.value or "").strip().lower()
        headers[val] = idx

    # Map expected columns (flexible matching)
    col_map = {}
    for key, patterns in {
        "office": ["office", "program", "agency"],
        "doc_number": ["document number", "doc number", "doc #", "doc_number", "number"],
        "title": ["title", "document title", "name"],
        "date": ["date", "issued", "effective date"],
        "type": ["type", "document type", "category"],
    }.items():
        for pat in patterns:
            for header_name, idx in headers.items():
                if pat in header_name:
                    col_map[key] = idx
                    break
            if key in col_map:
                break

    # All columns that might contain URLs
    url_cols = list(range(len(list(ws[1]))))

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        if row_idx < start_row:
            continue
        if end_row and row_idx > end_row:
            break

        # Skip completely empty rows
        if all(c.value is None for c in row):
            continue

        row_data = {
            "row_id": f"{row_idx:04d}",
            "office": str(row[col_map.get("office", 0)].value or "").strip(),
            "doc_number": str(row[col_map.get("doc_number", 1)].value or "").strip(),
            "title": str(row[col_map.get("title", 2)].value or "").strip(),
            "date": str(row[col_map.get("date", 3)].value or "").strip(),
            "doc_type": str(row[col_map.get("type", 4)].value or "").strip(),
            "urls": extract_urls_from_row(row, url_cols),
        }
        rows.append(row_data)

    wb.close()
    logger.info(f"Loaded {len(rows)} rows (range {start_row}-{end_row or 'end'})")
    return rows


# ── Portal cross-reference ───────────────────────────────────────────────

def load_portal_map(catalog_path):
    """
    Load the Master Catalog to get portal URLs and active/rescinded status.
    Returns dict mapping row_id -> {status, portal_url, ...}
    """
    logger = logging.getLogger("acf_v2")

    if not os.path.exists(catalog_path):
        logger.warning(f"Master Catalog not found at {catalog_path} — skipping portal source")
        return {}

    logger.info(f"Loading portal cross-reference from: {catalog_path}")
    wb = openpyxl.load_workbook(catalog_path, data_only=True)

    portal_map = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        headers = {}
        for idx, cell in enumerate(ws[1]):
            val = str(cell.value or "").strip().lower()
            headers[val] = idx

        # Look for row ID and URL columns
        id_col = None
        url_col = None
        status_col = None

        for name, idx in headers.items():
            if "row" in name and ("id" in name or "#" in name or "number" in name):
                id_col = idx
            elif name in ("id", "row_id", "row"):
                id_col = idx
            if "portal" in name and "url" in name:
                url_col = idx
            elif "url" in name or "link" in name:
                url_col = idx
            if "status" in name or "rescind" in name:
                status_col = idx

        if id_col is None:
            continue

        for row in ws.iter_rows(min_row=2):
            row_id = str(row[id_col].value or "").strip()
            if not row_id:
                continue
            row_id = f"{int(row_id):04d}" if row_id.isdigit() else row_id

            entry = {"sheet": ws_name}
            if url_col is not None:
                cell = row[url_col]
                # Check hyperlink
                if hasattr(cell, 'hyperlink') and cell.hyperlink and cell.hyperlink.target:
                    entry["portal_url"] = cell.hyperlink.target.strip()
                elif cell.value and str(cell.value).startswith("http"):
                    entry["portal_url"] = str(cell.value).strip()
            if status_col is not None:
                entry["status"] = str(row[status_col].value or "").strip()

            portal_map[row_id] = entry

    wb.close()
    logger.info(f"Loaded {len(portal_map)} portal cross-reference entries")
    return portal_map


# ── Source 1: HHS Guidance Portal ────────────────────────────────────────

def try_portal_source(row_data, portal_map, session, row_dir):
    """
    Try to download from the HHS Guidance Portal.
    Returns list of downloaded file dicts, or empty list.
    """
    logger = logging.getLogger("acf_v2")
    row_id = row_data["row_id"]

    portal_entry = portal_map.get(row_id, {})
    portal_url = portal_entry.get("portal_url", "")

    if not portal_url:
        return []

    logger.debug(f"  [PORTAL] Trying: {portal_url}")
    return try_download_url(portal_url, session, row_dir, source="portal")


# ── Source 2: Direct URLs from inventory ─────────────────────────────────

def try_direct_urls(row_data, session, row_dir):
    """
    Try each URL from the Excel inventory.
    Returns list of downloaded file dicts, or empty list.
    """
    logger = logging.getLogger("acf_v2")
    urls = row_data.get("urls", [])

    if not urls:
        return []

    all_files = []
    for url in urls:
        logger.debug(f"  [DIRECT] Trying: {url}")
        files = try_download_url(url, session, row_dir, source="direct")
        all_files.extend(files)
        if files:
            break  # Got something, stop trying
        time.sleep(REQUEST_DELAY)

    return all_files


# ── Source 3: Wayback Machine ────────────────────────────────────────────

def try_wayback_source(row_data, session, row_dir):
    """
    Try the Wayback Machine as last resort, with strict validation.
    Returns list of downloaded file dicts, or empty list.
    """
    logger = logging.getLogger("acf_v2")
    urls = row_data.get("urls", [])

    if not urls:
        return []

    for url in urls[:3]:  # Only try first 3 URLs to limit Wayback load
        wb_url = f"https://web.archive.org/web/2/{url}"
        logger.debug(f"  [WAYBACK] Trying: {wb_url}")

        try:
            time.sleep(WAYBACK_DELAY)
            resp = session.get(wb_url, timeout=30, allow_redirects=True)

            # Check if Wayback actually has it
            if resp.status_code != 200:
                continue

            final_url = resp.url

            # Reject if it redirected to a DOJ or error page
            parsed = urlparse(final_url)
            if "justice.gov" in parsed.netloc.lower():
                logger.debug(f"  [WAYBACK] Rejected — redirected to DOJ: {final_url}")
                continue

            content_type = resp.headers.get("Content-Type", "").lower()

            # If it's a direct file download from Wayback
            if any(ct in content_type for ct in ["pdf", "msword", "officedocument", "octet-stream"]):
                fname = guess_filename(resp, url)
                fpath = os.path.join(row_dir, fname)
                with open(fpath, "wb") as f:
                    f.write(resp.content)

                # Validate
                vr = validate_content(fpath)
                if vr.valid:
                    return [{
                        "filename": fname,
                        "source_url": url,
                        "wayback_url": wb_url,
                        "size_bytes": os.path.getsize(fpath),
                        "sha256": sha256_file(fpath),
                        "content_type": content_type,
                        "source": "wayback",
                        "validation": vr.to_dict(),
                    }]
                else:
                    logger.debug(f"  [WAYBACK] File failed validation: {vr.reason}")
                    os.remove(fpath)
                    continue

            # It's HTML — need to check content quality before accepting
            if "html" in content_type or "text" in content_type:
                html_content = resp.text

                # Quick pre-check before saving
                from validate import check_for_junk_indicators, DOJ_INDICATORS, WAYBACK_ERROR_INDICATORS
                doj = check_for_junk_indicators(html_content, DOJ_INDICATORS)
                if doj:
                    logger.debug(f"  [WAYBACK] Rejected — DOJ content detected")
                    continue
                wbe = check_for_junk_indicators(html_content, WAYBACK_ERROR_INDICATORS)
                if wbe:
                    logger.debug(f"  [WAYBACK] Rejected — Wayback error page")
                    continue

                # Save and validate
                html_fname = "wayback_page.html"
                html_path = os.path.join(row_dir, html_fname)
                with open(html_path, "w", encoding="utf-8") as f:
                    f.write(html_content)

                vr = validate_content(html_path)
                if vr.valid:
                    return [{
                        "filename": html_fname,
                        "source_url": url,
                        "wayback_url": wb_url,
                        "size_bytes": os.path.getsize(html_path),
                        "sha256": sha256_file(html_path),
                        "content_type": "text/html",
                        "source": "wayback",
                        "needs_conversion": True,
                        "validation": vr.to_dict(),
                    }]
                else:
                    logger.debug(f"  [WAYBACK] HTML failed validation: {vr.reason}")
                    os.remove(html_path)
                    continue

        except requests.exceptions.RequestException as e:
            logger.debug(f"  [WAYBACK] Request failed: {e}")
            continue

    return []


# ── Download helper ──────────────────────────────────────────────────────

def guess_filename(response, original_url):
    """Guess a good filename from the response or URL."""
    # Try Content-Disposition header
    cd = response.headers.get("Content-Disposition", "")
    if "filename=" in cd:
        match = re.search(r'filename[*]?="?([^";\n]+)"?', cd)
        if match:
            return safe_filename(match.group(1))

    # Try the final URL path
    parsed = urlparse(response.url)
    path_name = os.path.basename(parsed.path)
    if path_name and "." in path_name:
        return safe_filename(path_name)

    # Try original URL
    parsed = urlparse(original_url)
    path_name = os.path.basename(parsed.path)
    if path_name and "." in path_name:
        return safe_filename(path_name)

    # Guess extension from content type
    ct = response.headers.get("Content-Type", "").lower()
    ext_map = {
        "pdf": ".pdf", "msword": ".doc", "wordprocessingml": ".docx",
        "spreadsheetml": ".xlsx", "ms-excel": ".xls",
        "presentationml": ".pptx", "ms-powerpoint": ".ppt",
        "html": ".html", "plain": ".txt", "rtf": ".rtf",
    }
    ext = ".bin"
    for key, val in ext_map.items():
        if key in ct:
            ext = val
            break

    return f"document{ext}"


def try_download_url(url, session, row_dir, source="direct"):
    """
    Attempt to download from a URL, following redirects.
    Scrapes the page for downloadable file links if it's HTML.
    Returns list of downloaded file dicts.
    """
    logger = logging.getLogger("acf_v2")

    try:
        time.sleep(REQUEST_DELAY)
        resp = session.get(url, timeout=30, allow_redirects=True)
        resp.raise_for_status()
    except requests.exceptions.RequestException as e:
        logger.debug(f"  [{source.upper()}] Request failed for {url}: {e}")
        return []

    content_type = resp.headers.get("Content-Type", "").lower()

    # ── Direct file download ──
    if any(ct in content_type for ct in ["pdf", "msword", "officedocument",
                                          "octet-stream", "ms-excel",
                                          "ms-powerpoint", "rtf"]):
        fname = guess_filename(resp, url)
        fpath = os.path.join(row_dir, fname)

        # Avoid overwriting
        base, ext = os.path.splitext(fpath)
        counter = 1
        while os.path.exists(fpath):
            fpath = f"{base}_{counter}{ext}"
            counter += 1
        fname = os.path.basename(fpath)

        with open(fpath, "wb") as f:
            f.write(resp.content)

        # Validate
        vr = validate_content(fpath)
        if vr.valid:
            return [{
                "filename": fname,
                "source_url": url,
                "size_bytes": os.path.getsize(fpath),
                "sha256": sha256_file(fpath),
                "content_type": content_type,
                "source": source,
                "validation": vr.to_dict(),
            }]
        else:
            logger.debug(f"  [{source.upper()}] File failed validation: {vr.reason}")
            os.remove(fpath)
            return []

    # ── HTML page — scrape for file links ──
    if "html" in content_type or "text" in content_type:
        return scrape_page_for_files(resp, url, session, row_dir, source)

    return []


def scrape_page_for_files(response, page_url, session, row_dir, source):
    """
    Parse an HTML page and look for downloadable file links.
    If no files found, save the page content for Playwright conversion later.
    """
    logger = logging.getLogger("acf_v2")

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        logger.warning("BeautifulSoup not available — saving raw HTML")
        return save_html_page(response.text, page_url, row_dir, source)

    soup = BeautifulSoup(response.text, "html.parser")
    file_extensions = {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".rtf", ".txt"}

    downloaded = []
    seen_urls = set()

    # Find all links to downloadable files
    for tag in soup.find_all("a", href=True):
        href = tag["href"].strip()
        if not href:
            continue

        # Resolve relative URLs
        full_url = urljoin(page_url, href)
        parsed = urlparse(full_url)
        ext = os.path.splitext(parsed.path)[1].lower()

        if ext in file_extensions and full_url not in seen_urls:
            seen_urls.add(full_url)
            logger.debug(f"  [{source.upper()}] Found file link: {os.path.basename(parsed.path)}")

            try:
                time.sleep(REQUEST_DELAY)
                dl_resp = session.get(full_url, timeout=30, allow_redirects=True)
                dl_resp.raise_for_status()

                fname = guess_filename(dl_resp, full_url)
                fpath = os.path.join(row_dir, fname)

                base, fext = os.path.splitext(fpath)
                counter = 1
                while os.path.exists(fpath):
                    fpath = f"{base}_{counter}{fext}"
                    counter += 1
                fname = os.path.basename(fpath)

                with open(fpath, "wb") as f:
                    f.write(dl_resp.content)

                vr = validate_content(fpath)
                if vr.valid:
                    downloaded.append({
                        "filename": fname,
                        "source_url": full_url,
                        "page_url": page_url,
                        "size_bytes": os.path.getsize(fpath),
                        "sha256": sha256_file(fpath),
                        "content_type": dl_resp.headers.get("Content-Type", ""),
                        "source": source,
                        "download_type": "scraped_from_page",
                        "validation": vr.to_dict(),
                    })
                else:
                    logger.debug(f"  [{source.upper()}] Scraped file failed validation: {vr.reason}")
                    os.remove(fpath)

            except requests.exceptions.RequestException as e:
                logger.debug(f"  [{source.upper()}] Failed to download {full_url}: {e}")

    if downloaded:
        return downloaded

    # No downloadable files found — save the HTML page itself
    return save_html_page(response.text, page_url, row_dir, source)


def save_html_page(html_content, page_url, row_dir, source):
    """Save HTML page content, validate it, mark for conversion."""
    logger = logging.getLogger("acf_v2")

    html_path = os.path.join(row_dir, "page_content.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    vr = validate_content(html_path)
    if vr.valid:
        return [{
            "filename": "page_content.html",
            "source_url": page_url,
            "size_bytes": os.path.getsize(html_path),
            "sha256": sha256_file(html_path),
            "content_type": "text/html",
            "source": source,
            "needs_conversion": True,
            "validation": vr.to_dict(),
        }]
    else:
        logger.debug(f"  [{source.upper()}] Saved HTML failed validation: {vr.reason}")
        # Keep the file but mark it as failed so we can review
        return [{
            "filename": "page_content.html",
            "source_url": page_url,
            "size_bytes": os.path.getsize(html_path),
            "content_type": "text/html",
            "source": source,
            "validation_failed": True,
            "validation": vr.to_dict(),
        }]


# ── Orchestrator ─────────────────────────────────────────────────────────

def harvest_row(row_data, portal_map, session, output_dir, dry_run=False):
    """
    Harvest a single inventory row through the source chain.
    Returns metadata dict.
    """
    logger = logging.getLogger("acf_v2")
    row_id = row_data["row_id"]
    row_dir = os.path.join(output_dir, row_id)

    metadata = {
        "row_id": row_id,
        "office": row_data["office"],
        "doc_number": row_data["doc_number"],
        "title": row_data["title"],
        "date": row_data["date"],
        "doc_type": row_data["doc_type"],
        "urls": row_data["urls"],
        "portal_status": portal_map.get(row_id, {}).get("status", "unknown"),
        "status": "pending",
        "source_used": None,
        "files_downloaded": [],
        "attempts": [],
        "harvested_at": datetime.now(timezone.utc).isoformat(),
        "harvester_version": "2.0",
    }

    if dry_run:
        url_count = len(row_data["urls"])
        portal = "✓" if row_id in portal_map else "✗"
        logger.info(f"  [DRY] {row_id} | {row_data['office']:6s} | "
                     f"{url_count} URLs | Portal: {portal} | {row_data['doc_number'][:50]}")
        return metadata

    os.makedirs(row_dir, exist_ok=True)

    # ── Source chain ──

    # 1. Portal
    files = try_portal_source(row_data, portal_map, session, row_dir)
    metadata["attempts"].append({"source": "portal", "result": "success" if files else "no_match"})
    if files and not any(f.get("validation_failed") for f in files):
        metadata["files_downloaded"] = files
        metadata["source_used"] = "portal"
        metadata["status"] = "success"
        save_metadata(metadata, row_dir)
        return metadata

    # 2. Direct URLs
    files = try_direct_urls(row_data, session, row_dir)
    metadata["attempts"].append({"source": "direct", "result": "success" if files else "failed"})
    if files and not any(f.get("validation_failed") for f in files):
        metadata["files_downloaded"] = files
        metadata["source_used"] = "direct"
        metadata["status"] = "success"
        save_metadata(metadata, row_dir)
        return metadata

    # 3. Wayback Machine (with validation)
    files = try_wayback_source(row_data, session, row_dir)
    metadata["attempts"].append({"source": "wayback", "result": "success" if files else "failed"})
    if files and not any(f.get("validation_failed") for f in files):
        metadata["files_downloaded"] = files
        metadata["source_used"] = "wayback"
        metadata["status"] = "success"
        save_metadata(metadata, row_dir)
        return metadata

    # ── All sources exhausted ──
    # Check if we have any files at all (even validation-failed ones)
    all_files_in_dir = [f for f in os.listdir(row_dir) if f != "metadata.json"] if os.path.exists(row_dir) else []

    if files and any(f.get("validation_failed") for f in files):
        metadata["files_downloaded"] = files
        metadata["status"] = "validation_failed"
        metadata["source_used"] = files[0].get("source", "unknown")
    elif row_data["urls"]:
        metadata["status"] = "failed"
    else:
        metadata["status"] = "no_urls"

    save_metadata(metadata, row_dir)
    return metadata


def save_metadata(metadata, row_dir):
    """Save metadata.json to row directory."""
    meta_path = os.path.join(row_dir, "metadata.json")
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False, default=str)


# ── Validate-only mode ───────────────────────────────────────────────────

def validate_existing(output_dir):
    """Re-validate all files in an existing output directory."""
    logger = logging.getLogger("acf_v2")
    logger.info(f"Validating existing output: {output_dir}")

    stats = {
        "total_rows": 0, "valid_rows": 0, "invalid_rows": 0,
        "reasons": {},
    }
    invalid_details = []

    for folder_name in sorted(os.listdir(output_dir)):
        row_dir = os.path.join(output_dir, folder_name)
        if not os.path.isdir(row_dir) or not folder_name.isdigit():
            continue

        stats["total_rows"] += 1
        row_valid = False

        for fname in os.listdir(row_dir):
            if fname == "metadata.json":
                continue
            fpath = os.path.join(row_dir, fname)
            if not os.path.isfile(fpath):
                continue

            vr = validate_content(fpath)
            if vr.valid:
                row_valid = True
            else:
                reason = vr.reason
                stats["reasons"][reason] = stats["reasons"].get(reason, 0) + 1
                invalid_details.append({
                    "row": folder_name,
                    "file": fname,
                    "reason": reason,
                    "details": vr.details,
                })

        if row_valid:
            stats["valid_rows"] += 1
        else:
            stats["invalid_rows"] += 1

    # Report
    print("\n" + "=" * 60)
    print("VALIDATION REPORT")
    print("=" * 60)
    print(f"  Total rows:    {stats['total_rows']}")
    print(f"  Valid:         {stats['valid_rows']}")
    print(f"  Invalid:       {stats['invalid_rows']}")
    print(f"\n  Failure reasons:")
    for reason, count in sorted(stats["reasons"].items(), key=lambda x: -x[1]):
        print(f"    {reason:30s}  {count}")

    # Save report
    os.makedirs(REPORTS_DIR, exist_ok=True)
    report_path = os.path.join(REPORTS_DIR,
                               f"validation_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    with open(report_path, "w") as f:
        json.dump({"stats": stats, "invalid_details": invalid_details}, f, indent=2, default=str)
    print(f"\n  Report saved: {report_path}")

    return stats


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="ACF Guidance Harvester v2")
    parser.add_argument("--start", type=int, default=1, help="Start row (1-indexed)")
    parser.add_argument("--end", type=int, default=None, help="End row")
    parser.add_argument("--resume", action="store_true", help="Skip rows with existing output")
    parser.add_argument("--dry-run", action="store_true", help="Preview without downloading")
    parser.add_argument("--validate-only", action="store_true", help="Re-validate existing output_v2")
    parser.add_argument("--verbose", "-v", action="store_true", help="Debug logging")
    parser.add_argument("--input", type=str, default=None, help="Override input Excel path")
    parser.add_argument("--output", type=str, default=None, help="Override output directory")
    args = parser.parse_args()

    logger = setup_logging(args.verbose)

    output_dir = args.output or OUTPUT_DIR

    # Validate-only mode
    if args.validate_only:
        validate_existing(output_dir)
        return

    input_excel = args.input or INPUT_EXCEL

    if not os.path.exists(input_excel):
        logger.error(f"Input file not found: {input_excel}")
        logger.error("Use --input to specify the Excel file path")
        sys.exit(1)

    # Load data
    rows = load_inventory(input_excel, args.start, args.end)
    portal_map = load_portal_map(CATALOG_EXCEL)
    session = setup_session()

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(REPORTS_DIR, exist_ok=True)

    # Stats
    stats = {
        "total": len(rows),
        "success": 0,
        "failed": 0,
        "no_urls": 0,
        "validation_failed": 0,
        "skipped": 0,
        "sources": {"portal": 0, "direct": 0, "wayback": 0},
    }

    start_time = datetime.now()
    logger.info("=" * 60)
    logger.info("ACF GUIDANCE HARVESTER v2")
    logger.info("=" * 60)
    logger.info(f"  Rows to process: {len(rows)}")
    logger.info(f"  Portal entries:  {len(portal_map)}")
    logger.info(f"  Output dir:      {output_dir}")
    logger.info(f"  Resume mode:     {args.resume}")
    logger.info(f"  Dry run:         {args.dry_run}")
    logger.info("=" * 60)

    for i, row_data in enumerate(rows):
        row_id = row_data["row_id"]

        # Resume check
        if args.resume:
            meta_path = os.path.join(output_dir, row_id, "metadata.json")
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, "r") as f:
                        existing = json.load(f)
                    if existing.get("status") == "success":
                        stats["skipped"] += 1
                        continue
                except Exception:
                    pass

        # Progress
        pct = ((i + 1) / len(rows)) * 100
        logger.info(f"[{i+1}/{len(rows)} {pct:.0f}%] {row_id} | "
                     f"{row_data['office']:6s} | {row_data['doc_number'][:40]}")

        # Harvest
        try:
            metadata = harvest_row(row_data, portal_map, session, output_dir, args.dry_run)
        except Exception as e:
            logger.error(f"  UNEXPECTED ERROR: {e}")
            stats["failed"] += 1
            continue

        # Update stats
        status = metadata.get("status", "failed")
        if status == "success":
            stats["success"] += 1
            src = metadata.get("source_used", "unknown")
            if src in stats["sources"]:
                stats["sources"][src] += 1
        elif status == "no_urls":
            stats["no_urls"] += 1
        elif status == "validation_failed":
            stats["validation_failed"] += 1
        else:
            stats["failed"] += 1

    # ── Summary ──
    elapsed = datetime.now() - start_time
    logger.info("")
    logger.info("=" * 60)
    logger.info("HARVEST v2 COMPLETE")
    logger.info("=" * 60)
    logger.info(f"  Total rows:           {stats['total']}")
    logger.info(f"  Success:              {stats['success']}")
    logger.info(f"  Failed:               {stats['failed']}")
    logger.info(f"  Validation failed:    {stats['validation_failed']}")
    logger.info(f"  No URLs:              {stats['no_urls']}")
    logger.info(f"  Skipped (resume):     {stats['skipped']}")
    logger.info(f"  Sources used:")
    logger.info(f"    Portal:             {stats['sources']['portal']}")
    logger.info(f"    Direct:             {stats['sources']['direct']}")
    logger.info(f"    Wayback:            {stats['sources']['wayback']}")
    logger.info(f"  Elapsed:              {elapsed}")

    # Save report
    report = {
        "stats": stats,
        "started_at": start_time.isoformat(),
        "elapsed_seconds": elapsed.total_seconds(),
        "config": {
            "input": input_excel,
            "output": output_dir,
            "start_row": args.start,
            "end_row": args.end,
            "resume": args.resume,
        },
    }
    report_path = os.path.join(REPORTS_DIR,
                               f"harvest_v2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    with open(report_path, "w") as f:
        json.dump(report, f, indent=2, default=str)
    logger.info(f"  Report saved:         {report_path}")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
