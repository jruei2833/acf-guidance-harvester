#!/usr/bin/env python3
"""
ACF Guidance — Build Delivery Spreadsheet with S3 Links
=========================================================
Reads the Master Catalog and ACF_Guidance folder, matches files to rows,
and generates a clean Excel spreadsheet with clickable S3 Console links.

Usage:
  python build_spreadsheet.py
"""

import os
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if not os.path.exists(os.path.join(BASE_DIR, "dev")):
    BASE_DIR = os.path.dirname(BASE_DIR)

CATALOG_PATH = os.path.join(BASE_DIR, "dev", "catalog", "ACF_Guidance_Master_Catalog.xlsx")
ACF_DIR = os.path.join(BASE_DIR, "ACF_Guidance")
OUTPUT_PATH = os.path.join(BASE_DIR, "ACF_Guidance_Index.xlsx")

S3_BUCKET = "ofra-acf-guidance"
S3_REGION = "us-east-2"

# S3 Console link format
S3_CONSOLE_URL = (
    "https://us-east-2.console.aws.amazon.com/s3/object/{bucket}"
    "?region={region}&bucketType=general&prefix={key}"
)


def build_file_map(acf_dir):
    """
    Build a mapping from row ID to list of files in ACF_Guidance.
    Files are named like: 0002_AT-79-12.pdf, 0002_AT-79-12_attachment.pdf
    """
    file_map = defaultdict(list)
    if not os.path.isdir(acf_dir):
        print(f"ERROR: ACF_Guidance directory not found: {acf_dir}")
        return file_map

    for fname in os.listdir(acf_dir):
        fpath = os.path.join(acf_dir, fname)
        if not os.path.isfile(fpath):
            continue
        # Extract row ID from filename prefix (first 4 digits before underscore)
        match = re.match(r'^(\d{4})_', fname)
        if match:
            row_id = int(match.group(1))
            file_map[row_id].append({
                "filename": fname,
                "size_bytes": os.path.getsize(fpath),
                "extension": os.path.splitext(fname)[1].lower(),
            })

    return file_map


def s3_console_link(filename):
    """Build an S3 Console URL for a file."""
    encoded = quote(filename, safe='')
    return S3_CONSOLE_URL.format(
        bucket=S3_BUCKET, region=S3_REGION, key=encoded
    )


def build_spreadsheet():
    print(f"Loading catalog: {CATALOG_PATH}")
    wb_src = openpyxl.load_workbook(CATALOG_PATH, data_only=True)
    ws_src = wb_src.active

    print(f"Scanning files: {ACF_DIR}")
    file_map = build_file_map(ACF_DIR)
    print(f"  Mapped {sum(len(v) for v in file_map.values())} files across {len(file_map)} rows")

    # Create output workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ACF Guidance Index"

    # ── Styles ──
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B3A5C")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    date_font = Font(name="Arial", size=10)

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    alt_fill = PatternFill("solid", fgColor="F2F7FB")

    # ── Headers ──
    headers = [
        ("Row ID", 8),
        ("Office", 10),
        ("Document Type", 22),
        ("Document Number", 20),
        ("Title", 50),
        ("Issue Date", 14),
        ("File Name", 40),
        ("File Type", 10),
        ("File Size", 12),
        ("S3 Link", 14),
        ("# Files", 8),
        ("Source URL", 14),
    ]

    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:L1"

    # ── Data rows ──
    row_num = 2
    matched = 0
    unmatched = 0

    for src_row in ws_src.iter_rows(min_row=2, values_only=False):
        row_id = src_row[0].value  # Row ID (int)
        if row_id is None:
            continue

        row_id_int = int(row_id)
        office = str(src_row[1].value or "")
        doc_type = str(src_row[2].value or "")
        doc_number = str(src_row[3].value or "")
        title = str(src_row[5].value or "")  # col index 5 = Title
        issue_date = src_row[6].value  # Issue Date
        source_url = str(src_row[15].value or "")  # Source URL (col 16, index 15)

        files = file_map.get(row_id_int, [])

        if not files:
            # Row with no files — still include it
            fill = alt_fill if (row_num % 2 == 0) else None

            cells = [
                (f"{row_id_int:04d}", data_font),
                (office, data_font),
                (doc_type, data_font),
                (doc_number, data_font),
                (title, data_font),
                (issue_date, date_font),
                ("—", data_font),
                ("—", data_font),
                ("—", data_font),
                ("No file", data_font),
                (0, data_font),
            ]

            for col_idx, (value, font) in enumerate(cells, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = font
                cell.border = thin_border
                if fill:
                    cell.fill = fill

            # Source URL
            if source_url and source_url.startswith("http"):
                cell = ws.cell(row=row_num, column=12, value="Source")
                cell.hyperlink = source_url
                cell.font = link_font
            else:
                ws.cell(row=row_num, column=12, value="—").font = data_font

            ws.cell(row=row_num, column=12).border = thin_border

            row_num += 1
            unmatched += 1
            continue

        matched += 1

        # Primary file (first one, usually the main document)
        # Sort: PDFs first, then by name
        files.sort(key=lambda f: (0 if f["extension"] == ".pdf" else 1, f["filename"]))

        for file_idx, file_info in enumerate(files):
            fill = alt_fill if (row_num % 2 == 0) else None
            fname = file_info["filename"]
            ext = file_info["extension"].lstrip(".")
            size_kb = file_info["size_bytes"] / 1024

            if size_kb >= 1024:
                size_str = f"{size_kb/1024:.1f} MB"
            else:
                size_str = f"{size_kb:.0f} KB"

            # Only show metadata on first file row
            if file_idx == 0:
                cells = [
                    (f"{row_id_int:04d}", data_font),
                    (office, data_font),
                    (doc_type, data_font),
                    (doc_number, data_font),
                    (title, data_font),
                    (issue_date, date_font),
                ]
            else:
                cells = [
                    ("", data_font),
                    ("", data_font),
                    ("", data_font),
                    ("", data_font),
                    ("", data_font),
                    ("", data_font),
                ]

            for col_idx, (value, font) in enumerate(cells, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = font
                cell.border = thin_border
                if fill:
                    cell.fill = fill

            # File Name
            cell = ws.cell(row=row_num, column=7, value=fname)
            cell.font = data_font
            cell.border = thin_border
            if fill: cell.fill = fill

            # File Type
            cell = ws.cell(row=row_num, column=8, value=ext.upper())
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if fill: cell.fill = fill

            # File Size
            cell = ws.cell(row=row_num, column=9, value=size_str)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            if fill: cell.fill = fill

            # S3 Link
            s3_url = s3_console_link(fname)
            cell = ws.cell(row=row_num, column=10, value="Open in S3")
            cell.hyperlink = s3_url
            cell.font = link_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if fill: cell.fill = fill

            # File count (only on first row)
            if file_idx == 0:
                cell = ws.cell(row=row_num, column=11, value=len(files))
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if fill: cell.fill = fill
            else:
                cell = ws.cell(row=row_num, column=11, value="")
                cell.border = thin_border
                if fill: cell.fill = fill

            # Source URL (only on first row)
            if file_idx == 0 and source_url and source_url.startswith("http"):
                cell = ws.cell(row=row_num, column=12, value="Source")
                cell.hyperlink = source_url
                cell.font = link_font
            else:
                cell = ws.cell(row=row_num, column=12, value="" if file_idx > 0 else "—")
                cell.font = data_font

            ws.cell(row=row_num, column=12).border = thin_border
            if fill: ws.cell(row=row_num, column=12).fill = fill

            row_num += 1

    # ── Summary sheet ──
    ws_sum = wb.create_sheet("Summary")

    sum_headers = ["Metric", "Count"]
    for col, header in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 15

    total_files = sum(len(v) for v in file_map.values())
    summary_data = [
        ("Total catalog entries", ws_src.max_row - 1),
        ("Entries with files", matched),
        ("Entries without files", unmatched),
        ("Total files in S3", total_files),
        ("S3 Bucket", S3_BUCKET),
        ("S3 Region", S3_REGION),
    ]

    # Office breakdown
    office_counts = defaultdict(int)
    for src_row in ws_src.iter_rows(min_row=2, values_only=True):
        if src_row[0] is not None:
            office_counts[str(src_row[1] or "Unknown")] += 1

    for r_idx, (metric, value) in enumerate(summary_data, 2):
        ws_sum.cell(row=r_idx, column=1, value=metric).font = data_font
        cell = ws_sum.cell(row=r_idx, column=2, value=value)
        cell.font = data_font
        cell.alignment = Alignment(horizontal="right")

    # Office breakdown
    r_idx = len(summary_data) + 3
    ws_sum.cell(row=r_idx, column=1, value="Documents by Office").font = Font(name="Arial", size=11, bold=True)
    r_idx += 1
    for office, count in sorted(office_counts.items(), key=lambda x: -x[1]):
        ws_sum.cell(row=r_idx, column=1, value=office).font = data_font
        cell = ws_sum.cell(row=r_idx, column=2, value=count)
        cell.font = data_font
        cell.alignment = Alignment(horizontal="right")
        r_idx += 1

    # ── Save ──
    wb.save(OUTPUT_PATH)
    wb_src.close()

    print(f"\n{'='*60}")
    print(f"SPREADSHEET GENERATED")
    print(f"{'='*60}")
    print(f"  Output:          {OUTPUT_PATH}")
    print(f"  Catalog entries:  {ws_src.max_row - 1}")
    print(f"  With files:      {matched}")
    print(f"  Without files:   {unmatched}")
    print(f"  Total files:     {total_files}")
    print(f"  Data rows:       {row_num - 2}")
    print(f"{'='*60}")


if __name__ == "__main__":
    build_spreadsheet()
